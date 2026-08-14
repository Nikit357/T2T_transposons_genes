#!/usr/bin/env python
"""17_build_supplementary_260809.py — the 2026-08-09 supplementary package.

Builds `supplementary_260809/` from `supplementary/`, applying the changes the
2026-08-09 scientific review requires. It never edits the 2026-08-05 package in place.

What changes and why
--------------------
1. **File S3 gains a `GO_borderline` sheet.** The Results quote FDR values for terms in the
   0.05 < FDR <= 0.1 band (flavone metabolic process at 0.088, glutamatergic synapse at 0.086,
   the MIR metal terms at 0.078 ...). Those terms are *excluded* from every shipped sheet,
   because every shipped GO sheet is cut at FDR < 0.05. A number the main text argues from may
   not live outside the supplementary package (principle P2), so the band is now shipped.

2. **File S4 sheets are reordered to the order the Results cite them**, and the file is renamed
   to match: interferon-alpha domain (Results), then the prior-work comparison (Results), then
   the assembly bound (Discussion). `assembly_bound` previously sat between the two Results
   subjects (principle P3).

3. **File S5 gains no data**; `headline_by_condition` moves next to the percentile sheets it
   belongs with, which is where the rewritten sensitivity subsection cites it.

4. README, INVENTORY.json and CHECKSUMS.sha256 are regenerated so they describe what is
   actually in the folder.

File *numbers* do not change: after the 2026-08-09 rewrite adds File S4 citations to the
interferon-alpha and prior-work subsections, S1..S6 are already cited in ascending order.

Run:  ~/venvs/Retroelements_3_11/bin/python revision_G3/17_build_supplementary_260809.py
      ~/venvs/Retroelements_3_11/bin/python revision_G3/17_build_supplementary_260809.py --verify
"""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
SRC = HERE / "supplementary"
DST = HERE / "supplementary_260809"
OUT = HERE / "output"

# FDR band that the Results discuss but no shipped sheet contains
BAND_LO, BAND_HI = 0.05, 0.10

# reference tables at FDR 0.1, from which the band is cut
BORDERLINE_SOURCES = [
    ("classes_by_count", OUT / "GO_classes_count_fdr01_reference.csv", "class_name"),
    ("classes_by_divergence", OUT / "GO_classes_divergence_fdr01_reference.csv", "class_name"),
    ("families", OUT / "GO_families_fdr01_reference.csv", "family_name"),
]

# File S4: the order the Results and Discussion cite these sheets
S4_SHEET_ORDER = [
    "README",
    "IFNA_elements",
    "IFNA_tests",
    "IFNA_subfamily_composition",
    "prior_work_overlap_matrix",
    "prior_work_categories",
    "prior_work_shared_genes",
    "assembly_bound",
]

S5_SHEET_ORDER = [
    "README",
    "window_classes",
    "window_families",
    "window_concordance",
    "window_flips",
    "geneset_stability",
    "rank_stability",
    "percentile_summary",
    "percentile_terms",
    "headline_by_condition",
    "GO_grid_index",
    "GO_grid_preservation",
    "GO_grid_concordance",
    "GO_grid_terms",
]

RENAMES = {
    "File_S4_IFNA_domain_and_prior_work.xlsx":
        "File_S4_IFNA_domain_prior_work_and_assembly.xlsx",
}


def build_borderline() -> pd.DataFrame:
    """Every GO term with 0.05 < FDR <= 0.1, at all three levels of analysis."""
    frames = []
    for level, path, group_col in BORDERLINE_SOURCES:
        if not path.exists():
            raise SystemExit(f"missing reference table: {path}")
        d = pd.read_csv(path)
        band = d[(d["FDR"] > BAND_LO) & (d["FDR"] <= BAND_HI)].copy()
        band.insert(0, "level", level)
        band = band.rename(columns={group_col: "group"})
        keep = ["level", "group", "Term ID", "Term Name", "Term Database",
                "P-value", "FDR", "Fold Enrichment", "Overlap Count",
                "Total Term Genes (Human)", "Overlapping Genes"]
        band = band[[c for c in keep if c in band.columns]]
        frames.append(band.sort_values(["group", "FDR"]))
    out = pd.concat(frames, ignore_index=True)
    out["FDR"] = out["FDR"].round(6)
    out["P-value"] = out["P-value"].astype(float)
    return out


def write_sheet(xlsx: Path, name: str, df: pd.DataFrame, position: int | None = None):
    wb = openpyxl.load_workbook(xlsx)
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(["" if pd.isna(v) else v for v in row])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    ws.freeze_panes = "A2"
    if position is not None:
        wb.move_sheet(name, offset=position - wb.sheetnames.index(name))
    wb.save(xlsx)


def reorder_sheets(xlsx: Path, order: list[str]):
    wb = openpyxl.load_workbook(xlsx)
    present = [s for s in order if s in wb.sheetnames]
    missing = [s for s in wb.sheetnames if s not in order]
    if missing:
        raise SystemExit(f"{xlsx.name}: sheets not in the declared order: {missing}")
    for target, name in enumerate(present):
        wb.move_sheet(name, offset=target - wb.sheetnames.index(name))
    wb.save(xlsx)
    return present


def patch_readme(xlsx: Path, additions: dict[str, str]):
    """Append/replace rows in the workbook's README sheet."""
    wb = openpyxl.load_workbook(xlsx)
    if "README" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["README"]
    existing = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            existing[str(v)] = r
    for sheet, desc in additions.items():
        if sheet in existing:
            ws.cell(existing[sheet], 2, desc)
        else:
            ws.append([sheet, desc, "", ""])
    wb.save(xlsx)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def build():
    if not SRC.exists():
        raise SystemExit(f"source package not found: {SRC}")
    DST.mkdir(exist_ok=True)

    # 1. copy every workbook, applying the renames
    copied = {}
    for src in sorted(SRC.glob("*.xlsx")):
        dst = DST / RENAMES.get(src.name, src.name)
        shutil.copy2(src, dst)
        copied[src.name] = dst
        print(f"  copied  {src.name} -> {dst.name}")

    # 2. File S3: ship the borderline band
    s3 = copied["File_S3_gene_ontology.xlsx"]
    band = build_borderline()
    write_sheet(s3, "GO_borderline", band)
    patch_readme(s3, {
        "GO_borderline":
            f"GO terms with {BAND_LO} < FDR <= {BAND_HI}, i.e. those the Results name as "
            f"narrowly not significant. Not part of any reported result; supplied so that "
            f"every FDR value quoted in the text can be checked. {len(band)} rows.",
    })
    print(f"  File S3: added GO_borderline, {len(band)} rows "
          f"({band.level.value_counts().to_dict()})")

    # 3. File S4: reorder to citation order
    s4 = copied["File_S4_IFNA_domain_and_prior_work.xlsx"]
    order4 = reorder_sheets(s4, S4_SHEET_ORDER)
    print(f"  File S4: sheet order -> {order4}")

    # 4. File S5: headline_by_condition next to the percentile sheets
    s5 = copied["File_S5_sensitivity_and_robustness.xlsx"]
    order5 = reorder_sheets(s5, S5_SHEET_ORDER)
    print(f"  File S5: sheet order -> {order5}")

    # 5. regenerate the manifest and checksums
    inventory = {
        "package": "supplementary_260809",
        "built_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "built_by": "revision_G3/17_build_supplementary_260809.py",
        "source_package": "supplementary/ (2026-08-05)",
        "changes": [
            "File S3: new sheet GO_borderline (0.05 < FDR <= 0.1) so every FDR quoted in the "
            "Results is checkable in the package",
            "File S4: sheets reordered to the order the text cites them; renamed to "
            "File_S4_IFNA_domain_prior_work_and_assembly.xlsx",
            "File S5: headline_by_condition moved next to the percentile sheets",
            "File numbers S1-S6 unchanged: already cited in ascending order",
        ],
        "files": [],
    }
    for f in sorted(DST.glob("*.xlsx")):
        wb = openpyxl.load_workbook(f, read_only=True)
        sheets = []
        for ws in wb.worksheets:
            sheets.append({"name": ws.title,
                           "rows": max(ws.max_row - 1, 0),
                           "cols": ws.max_column})
        wb.close()
        inventory["files"].append({
            "name": f.name,
            "bytes": f.stat().st_size,
            "sha256": sha256(f),
            "sheets": sheets,
        })
    (DST / "INVENTORY.json").write_text(json.dumps(inventory, indent=2) + "\n")

    lines = [f"{sha256(f)}  {f.name}" for f in sorted(DST.glob("*.xlsx"))]
    (DST / "CHECKSUMS.sha256").write_text("\n".join(lines) + "\n")

    total = sum(f.stat().st_size for f in DST.glob("*.xlsx"))
    print(f"\n  {len(inventory['files'])} workbooks, {total/1e6:.1f} MB -> {DST}")
    return inventory


def verify():
    inv = json.loads((DST / "INVENTORY.json").read_text())
    ok = True
    for entry in inv["files"]:
        f = DST / entry["name"]
        if not f.exists():
            print(f"  MISSING {entry['name']}"); ok = False; continue
        got = sha256(f)
        status = "ok" if got == entry["sha256"] else "CHECKSUM MISMATCH"
        if got != entry["sha256"]:
            ok = False
        print(f"  {status:<18} {entry['name']}  "
              f"{len(entry['sheets'])} sheets, {entry['bytes']/1e6:.2f} MB")
    # the band must be non-empty and must contain the terms the Results name
    band = pd.read_excel(DST / "File_S3_gene_ontology.xlsx", sheet_name="GO_borderline")
    for term in ["flavone metabolic process", "glutamatergic synapse",
                 "cellular response to cadmium ion", "nucleotide binding"]:
        hit = (band["Term Name"] == term).sum()
        print(f"  {'ok' if hit else 'MISSING':<18} GO_borderline contains {term!r} ({hit} rows)")
        if not hit:
            ok = False
    print("\nverify:", "PASS" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--verify", action="store_true")
    a = ap.parse_args()
    if a.verify:
        raise SystemExit(verify())
    build()
