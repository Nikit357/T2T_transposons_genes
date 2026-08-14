#!/usr/bin/env python
"""WP4 — gene-set overlap with Lu et al. 2020 (reviewer major comment 4, D6 option a).

The reviewer's comment
----------------------
"The author notes methodological differences... but does not directly compare
results on the same dataset. Is this difference due to methodology or the updated
genome assembly?"

Reference 14 = Lu JY, Shao W, Chang L, et al. Genomic Repeats Categorize Genes with
Distinct Functions for Orchestrated Regulation. Cell Rep 2020;30(10):3296-3311.e5.
DOI 10.1016/j.celrep.2020.02.048.

Decision D6 is option (a) only: compare at the level of gene sets, and state in the
response letter that a full reimplementation of their region-binning and clustering
was out of scope. This script is that comparison.

The species problem, which the plan did not anticipate
------------------------------------------------------
Their Table S1 categories are **mouse** gene sets, defined on mm9 with mouse SINE
subfamilies (B1, B2, B4) — not human ones:

    L1-enriched genes                    1,480
    Low-complexity-repeat-enriched genes 2,439
    SINE-enriched genes                  2,041

(A fourth category, 383 satellite-repeat-enriched genes, is described in their text
but is not in Table S1, so it cannot be compared.)

So "the same dataset" was never available in the sense the reviewer implies: any
comparison has to cross a species boundary via orthology, which is a real and
reportable limitation rather than a choice we made. Mouse symbols are mapped to
human symbols through the MGI mouse-human homology classes, and the mapping rate is
itself a result worth quoting: their L1-enriched category maps far worse than the
other two, because it is dominated by rodent-expanded families (Zfp*, Gm*, Trav*,
Cyp2b*) that have no single human counterpart.

Statistical universe
--------------------
Every test is restricted to genes that could in principle appear in both studies:
our background (the genes with a TSS window in `TEs_on_genes.csv`) intersected with
the set of human genes having a mouse ortholog. A gene with no mouse ortholog can
never be in one of their categories, so leaving it in the population would inflate
every odds ratio. Both sides are restricted to the same universe before testing.

Our sets are rebuilt by the same construction as the published GO foregrounds,
reused from `06_go_rerun_fdr005.py` so there is no second definition of "top 5 %".

Inputs (both downloaded, both recorded in external/PROVENANCE.md)
    external/lu2020/mmc2.xlsx            their Table S1
    external/HOM_MouseHumanSequence.rpt  MGI mouse-human homology classes

Outputs
    output/lu2020_categories_mapped.csv  every category gene, mouse -> human, path taken
    output/lu2020_mapping_summary.csv    mapping rate per category
    output/lu2020_overlap_matrix.csv     our set x their category: Fisher, Jaccard, overlap
    output/lu2020_overlap_genes.csv      the shared genes per pair, for inspection and
                                         for the supervenn panel built in Phase 4

Usage
    python revision_G3/04a_lu2020_geneset_overlap.py
    python revision_G3/04a_lu2020_geneset_overlap.py --summary
"""

from __future__ import annotations

import argparse
import importlib
import os
import sys

import numpy as np
import openpyxl
import pandas as pd
from scipy import stats
from statsmodels.stats.multitest import multipletests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import revision_lib as rl  # noqa: E402

go_rerun = importlib.import_module("06_go_rerun_fdr005")

OUTPUT_DIR = rl.OUTPUT_DIR
EXTERNAL_DIR = os.path.join(rl.REVISION_DIR, "external")
LU_TABLE_S1 = os.path.join(EXTERNAL_DIR, "lu2020", "mmc2.xlsx")
MGI_HOMOLOGY = os.path.join(EXTERNAL_DIR, "HOM_MouseHumanSequence.rpt")

# Counts stated in their Results text, used to verify we parsed the right columns.
EXPECTED_CATEGORY_SIZES = {
    "L1-enriched genes": 1480,
    "Low-complexity-repeat-enriched genes": 2439,
    "SINE-enriched genes": 2041,
}

# The three families that are the direct counterparts of their categories: L1 is
# their L1, Alu and MIR are the human SINEs their B1/B2/B4 correspond to.
COUNTERPART_FAMILIES = ["L1", "Alu", "MIR"]

CLASS_GROUPS = ["LINE", "LTR", "SINE", "DNA", "SVA", "Helitron", "TE_top", "TE_bottom"]


def load_lu_categories() -> dict[str, list[str]]:
    """Their Table S1: one column of mouse gene symbols per category."""
    workbook = openpyxl.load_workbook(LU_TABLE_S1, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    categories = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        genes = [
            str(row[index]).strip()
            for row in rows[1:]
            if index < len(row) and row[index] not in (None, "")
        ]
        # Their header has a double space; normalise so the expected-size check keys match.
        categories[" ".join(header.split())] = sorted(set(genes))

    for name, genes in categories.items():
        expected = EXPECTED_CATEGORY_SIZES.get(name)
        flag = ""
        if expected is not None:
            flag = " (matches their text)" if len(genes) == expected else \
                   f" (their text says {expected:,} — CHECK)"
        print(f"  {name:40s} {len(genes):>6,} mouse symbols{flag}")
    return categories


def load_ortholog_map() -> tuple[dict[str, str], set[str]]:
    """Mouse symbol -> human symbol, from MGI homology classes.

    Only mouse symbols with exactly one human symbol in their homology class are
    used. Ambiguous classes (rodent-expanded families, mostly) are dropped rather
    than resolved arbitrarily — an arbitrary pick would silently invent overlap.
    """
    homology = pd.read_csv(MGI_HOMOLOGY, sep="\t", low_memory=False)
    homology.columns = [c.strip() for c in homology.columns]

    mouse = homology[homology["Common Organism Name"] == "mouse, laboratory"]
    human = homology[homology["Common Organism Name"] == "human"]
    pairs = mouse[["DB Class Key", "Symbol"]].merge(
        human[["DB Class Key", "Symbol"]], on="DB Class Key",
        suffixes=("_mouse", "_human"),
    )

    counts = pairs.groupby("Symbol_mouse")["Symbol_human"].nunique()
    unique_mouse = set(counts[counts == 1].index)
    mapping = {
        row.Symbol_mouse: row.Symbol_human
        for row in pairs.itertuples()
        if row.Symbol_mouse in unique_mouse
    }
    human_with_ortholog = set(pairs["Symbol_human"])

    print(f"  MGI homology: {len(pairs):,} mouse-human pairs, "
          f"{len(mapping):,} mouse symbols with exactly one human ortholog, "
          f"{len(human_with_ortholog):,} human symbols with any mouse ortholog")
    return mapping, human_with_ortholog


def map_categories(
    categories: dict[str, list[str]],
    mapping: dict[str, str],
    background: set[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Map each category to human symbols, recording which path each gene took.

    Two paths, kept separate so the primary analysis never depends on the weaker one:
      `mgi_ortholog`  — an unambiguous MGI homology class (the primary path)
      `symbol_upcase` — no MGI class, but the uppercased mouse symbol is a gene in
                        our background. Standard practice for mouse/human symbol
                        pairs, but it is a guess, so it is reported and excluded
                        from the primary numbers.
    """
    records = []
    for category, genes in categories.items():
        for mouse_symbol in genes:
            human_symbol = mapping.get(mouse_symbol)
            if human_symbol is not None:
                path = "mgi_ortholog"
            elif mouse_symbol.upper() in background:
                human_symbol, path = mouse_symbol.upper(), "symbol_upcase"
            else:
                human_symbol, path = None, "unmapped"
            records.append(
                {
                    "category": category,
                    "mouse_symbol": mouse_symbol,
                    "human_symbol": human_symbol,
                    "mapping_path": path,
                    "in_our_background": human_symbol in background
                    if human_symbol else False,
                }
            )

    mapped = pd.DataFrame(records)
    summary = (
        mapped.assign(n=1)
        .pivot_table(index="category", columns="mapping_path", values="n",
                     aggfunc="sum", fill_value=0)
        .reset_index()
    )
    summary["n_mouse_genes"] = summary.get("mgi_ortholog", 0) + \
        summary.get("symbol_upcase", 0) + summary.get("unmapped", 0)
    summary["mgi_mapping_rate"] = summary.get("mgi_ortholog", 0) / summary["n_mouse_genes"]
    summary["n_testable"] = [
        int(((mapped.category == c) & (mapped.mapping_path == "mgi_ortholog")
             & mapped.in_our_background).sum())
        for c in summary["category"]
    ]
    return mapped, summary


def our_gene_sets(df: pd.DataFrame) -> dict[str, set[str]]:
    """The published top-5 % foreground sets, per class and for three families.

    Same constructions as the GO analysis (frozen cells 20 and 149, reused through
    `06_go_rerun_fdr005.py`), so this compares Lu et al. against exactly the sets the
    manuscript's GO results are built on.
    """
    top_n = go_rerun.CLASS_TOP_N
    sets: dict[str, set[str]] = {}

    for class_name in go_rerun.MAJOR_CLASSES:
        sets[class_name] = set(
            list(
                df.sort_values(f"{class_name}_number", ascending=False)[["Gene_name"]]
                .drop_duplicates()["Gene_name"]
            )[:top_n]
        )

    for class_name, column in [("SVA", "Retroposon_number"), ("Helitron", "RC_number")]:
        sets[class_name] = set(df[df[column] != 0]["Gene_name"].unique())

    for class_name, ascending in [("TE_top", False), ("TE_bottom", True)]:
        sets[class_name] = set(
            list(
                df.sort_values("TE_number", ascending=ascending)[["Gene_name"]]
                .drop_duplicates()["Gene_name"]
            )[:top_n]
        )

    take = df["Gene_name"].nunique() // 20
    for family_name in COUNTERPART_FAMILIES:
        column = f"{family_name}_number"
        genes = list(
            df[df[column] != 0]
            .sort_values(column, ascending=False)[["Gene_name"]]
            .drop_duplicates()["Gene_name"]
        )
        sets[f"{family_name} (family)"] = set(genes[:take] if len(genes) >= take else genes)

    return sets


def overlap_tests(
    ours: dict[str, set[str]],
    theirs: dict[str, set[str]],
    universe: set[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Fisher, Jaccard and overlap coefficient for every (our set, their category) pair."""
    rows, gene_rows = [], []
    for our_name, our_set in ours.items():
        a_set = our_set & universe
        for their_name, their_set in theirs.items():
            b_set = their_set & universe
            shared = a_set & b_set

            both = len(shared)
            only_a = len(a_set) - both
            only_b = len(b_set) - both
            neither = len(universe) - both - only_a - only_b
            odds_ratio, p_value = stats.fisher_exact(
                [[both, only_a], [only_b, neither]], alternative="two-sided"
            )
            expected = len(a_set) * len(b_set) / len(universe) if universe else np.nan

            rows.append(
                {
                    "our_set": our_name,
                    "their_category": their_name,
                    "n_ours_in_universe": len(a_set),
                    "n_theirs_in_universe": len(b_set),
                    "n_shared": both,
                    "n_expected_by_chance": expected,
                    "fold_over_expected": both / expected if expected else np.nan,
                    "jaccard": both / len(a_set | b_set) if (a_set | b_set) else np.nan,
                    "overlap_coefficient": both / min(len(a_set), len(b_set))
                    if a_set and b_set else np.nan,
                    "fisher_odds_ratio": odds_ratio,
                    "fisher_p_raw": p_value,
                }
            )
            for gene in sorted(shared):
                gene_rows.append(
                    {"our_set": our_name, "their_category": their_name, "gene": gene}
                )

    matrix = pd.DataFrame(rows)
    matrix["fisher_p_adjusted"] = multipletests(
        matrix["fisher_p_raw"], method="fdr_bh"
    )[1]
    matrix["significant"] = matrix["fisher_p_adjusted"] < rl.FDR_THRESHOLD
    matrix["direction"] = np.where(
        matrix["fisher_odds_ratio"] > 1, "excess", "depletion"
    )
    return matrix, pd.DataFrame(gene_rows)


def report() -> int:
    matrix_path = os.path.join(OUTPUT_DIR, "lu2020_overlap_matrix.csv")
    summary_path = os.path.join(OUTPUT_DIR, "lu2020_mapping_summary.csv")
    if not os.path.exists(matrix_path):
        print("Run without --summary first.", file=sys.stderr)
        return 1

    if os.path.exists(summary_path):
        print("=" * 78)
        print("Mouse -> human mapping of the Lu et al. 2020 categories")
        print("=" * 78)
        summary = pd.read_csv(summary_path)
        for r in summary.itertuples():
            print(f"  {r.category:40s} {r.n_mouse_genes:>6,} mouse -> "
                  f"{r.n_testable:>6,} testable human genes "
                  f"({100 * r.mgi_mapping_rate:.0f} % had a unique MGI ortholog)")

    matrix = pd.read_csv(matrix_path)
    print("\n" + "=" * 78)
    print(f"Gene-set overlap, fold over chance (FDR < {rl.FDR_THRESHOLD} in bold-equivalent *)")
    print("=" * 78)
    pivot = matrix.pivot(index="our_set", columns="their_category",
                         values="fold_over_expected")
    flags = matrix.pivot(index="our_set", columns="their_category", values="significant")
    header = "".join(f"{c[:26]:>28s}" for c in pivot.columns)
    print(f"{'our set':22s}{header}")
    for our_set in pivot.index:
        cells = "".join(
            f"{pivot.loc[our_set, c]:>26.2f}{'*' if flags.loc[our_set, c] else ' '} "
            for c in pivot.columns
        )
        print(f"{our_set:22s}{cells}")

    print("\nStrongest excesses (FDR < 0.05, ranked by fold over chance):")
    top = matrix[matrix.significant & (matrix.fold_over_expected > 1)].nlargest(
        8, "fold_over_expected"
    )
    for r in top.itertuples():
        print(f"  {r.our_set:22s} x {r.their_category:40s} "
              f"{r.fold_over_expected:.2f}x  OR = {r.fisher_odds_ratio:.2f}  "
              f"n = {r.n_shared:,}  FDR = {r.fisher_p_adjusted:.3g}")

    print("\nSignificant depletions:")
    dep = matrix[matrix.significant & (matrix.fold_over_expected < 1)].nsmallest(
        6, "fold_over_expected"
    )
    for r in dep.itertuples():
        print(f"  {r.our_set:22s} x {r.their_category:40s} "
              f"{r.fold_over_expected:.2f}x  OR = {r.fisher_odds_ratio:.2f}  "
              f"n = {r.n_shared:,}  FDR = {r.fisher_p_adjusted:.3g}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--summary", action="store_true",
                        help="print the comparison from existing outputs and exit")
    args = parser.parse_args()

    if args.summary:
        return report()

    for path in (LU_TABLE_S1, MGI_HOMOLOGY):
        if not os.path.exists(path):
            print(f"Missing input: {path}\nSee revision_G3/external/PROVENANCE.md",
                  file=sys.stderr)
            return 1

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Their Table S1 categories:")
    categories = load_lu_categories()

    print("\nOrtholog mapping:")
    mapping, human_with_ortholog = load_ortholog_map()

    df = go_rerun.load_gene_table()
    df = go_rerun.add_family_counts(df)
    background = set(df["Gene_name"].unique())

    mapped, summary = map_categories(categories, mapping, background)
    mapped.to_csv(os.path.join(OUTPUT_DIR, "lu2020_categories_mapped.csv"), index=False)
    summary.to_csv(os.path.join(OUTPUT_DIR, "lu2020_mapping_summary.csv"), index=False)

    # The primary analysis uses the MGI path only; the upcase rescue is recorded in
    # lu2020_categories_mapped.csv but never enters a test.
    primary = mapped[
        (mapped.mapping_path == "mgi_ortholog") & mapped.in_our_background
    ]
    theirs = {
        category: set(part["human_symbol"])
        for category, part in primary.groupby("category")
    }

    universe = background & human_with_ortholog
    print(f"\nTesting universe: {len(universe):,} genes "
          f"(our {len(background):,} background genes that have a mouse ortholog)")

    ours = our_gene_sets(df)
    matrix, shared_genes = overlap_tests(ours, theirs, universe)
    matrix.to_csv(os.path.join(OUTPUT_DIR, "lu2020_overlap_matrix.csv"), index=False)
    shared_genes.to_csv(
        os.path.join(OUTPUT_DIR, "lu2020_overlap_genes.csv"), index=False
    )
    print("Wrote lu2020_{categories_mapped,mapping_summary,overlap_matrix,overlap_genes}.csv")

    return report()


if __name__ == "__main__":
    sys.exit(main())
